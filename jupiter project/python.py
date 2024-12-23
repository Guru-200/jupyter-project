
import email
from email.header import decode_header
from email.message import EmailMessage
from email.mime.application import MIMEApplication
import imaplib
import fitz  
import openpyxl
from pymongo import MongoClient
from bson.binary import Binary  
import pandas as pd
from fpdf import FPDF
from PIL import Image
import io
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import smtplib
import re
import os
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
client = MongoClient("mongodb://localhost:27017/")
db = client["database"]
def properties(file_path):
    pdf = fitz.open(file_path)
    properties={}
    for page in pdf:
        text = page.get_text("text") # type: ignore
        lines = text.split("\n")

        
        for line in lines:
            if "Style number:" in line:    
                properties["Style number"] = line.split(":")[1].strip()
            if "Style:" in line and "Style number:" not in line:
                properties["Style"] = line.split(":")[1].strip()
            if "Brand:" in line:
                properties["Brand"] = line.split(":")[1].strip()
            if "Size:" in line:
                properties["Sizes"] = line.split(":")[1].strip()
            if "Commodity:" in line:
                properties["Commodity"] = line.split(":")[1].strip()
            if "Email:" in line:
                properties["E-mail"] = line.split(":")[1].strip()
            if "Care Address:" in line:
                properties["Care Address"] = line.split(":")[1].strip()

        
        if "Main image:" in text:
            image_refs = page.get_images(full=True)
            if image_refs:
                xref = image_refs[0][0]  
                image = pdf.extract_image(xref)
                image_bytes = image["image"]  
                properties["Main image"] = image_bytes
    pdf.close()
    return properties
def clean_newlines(text):
    """
    Replace \n between letters with a space.
    """
    if isinstance(text, str):
        return re.sub(r'(\w)\n(\w)', r'\1\2', text)  
    return text

def Tables(file_path):
    tables_list = []
    with pdfplumber.open(file_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
        
            tables = page.extract_tables()
        
        
            if tables:
                print(f"Tables found on page {page_number}:")
                for table_index, table in enumerate(tables, start=1):
                    print(f"Table {table_index}:")
                    if(page_number==2 and table_index==1):
                        index = pd.MultiIndex.from_tuples([
                        ('Dim', ''),
                        ('Description', ''),
                        ('Comment', ''),
                        ('Tol(-)', ''),
                        ('Tol(+)', ''),
                        ('XS', 'Increment'),
                        ('XS', 'Sample'),
                        ('XS', 'Deviation'),
                        ('S', 'Increment'),
                        ('S', 'Sample'),
                        ('S', 'Deviation')
                    ])
                        clean_table = [[clean_newlines(cell) for cell in row] for row in table]
                
                        df = pd.DataFrame(clean_table[2:], columns=index)
                        print(df)
                        tables_list.append(df)
                    elif(page_number==2 and table_index==2):
                        index = pd.MultiIndex.from_tuples([
                        ('M', 'Increment'),
                        ('M', 'Sample'),
                        ('M', 'Deviation'),
                        ('L', 'Increment'),
                        ('L', 'Sample'),
                        ('L', 'Deviation'),
                        ('XL', 'Increment'),
                        ('XL', 'Sample'),
                        ('XL', 'Deviation')
                    ])
                        clean_table = [[clean_newlines(cell) for cell in row] for row in table]
                        df = pd.DataFrame(clean_table[2:], columns=index)
                        print(df)
                        tables_list.append(df)
                    elif(page_number==3 and table_index==1):
                        clean_table = [[clean_newlines(cell) for cell in row] for row in table]
                        df = pd.DataFrame(clean_table[1:], columns=table[0])
                        print(df)
                        tables_list.append(df)
                    
            else:
                print(f"No tables found on page {page_number}.")

        return tables_list
def extract_sample_images_and_store_into_mongodb(file_path):
    doc = fitz.open(file_path)
    img_count = 0
    
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)  
        image_list = page.get_images(full=True)  
        if(page_num==0 or page_num==1):
            continue
        
        for img_index, img in enumerate(image_list):
            xref = img[0]  
            base_image = doc.extract_image(xref)
            image_bytes = base_image["image"]  
            img_type = base_image["ext"]  

            

            
            img_data = Binary(image_bytes)  

            
            image_doc = {
                'image': img_data,
                'image_type': img_type,
                'page': page_num + 1,
                'img_index': img_index + 1,
            }

            
            db.samples.insert_one(image_doc)
            img_count += 1

    print(f"Extracted {img_count} images and stored in MongoDB.")
def storeintomongodb(properties,tables):
    if properties:
        db.properties.insert_one(properties)
    print(tables)
    if tables:
        for i, table_df in enumerate(tables):
    
            print(i,table_df)
            print(table_df.columns)
            if isinstance(table_df.columns, pd.MultiIndex):
            
                table_df.columns = ['_'.join(col).strip() for col in table_df.columns]
            
            else:
                table_df.columns = [col if col is not None else "" for col in table_df.columns]
                
            print(table_df)
        
            table_dict = table_df.to_dict(orient="records")
            print(table_dict)
            print(pd.DataFrame(table_dict))
            document = {
                "table_name": f"Table_{i+1}",
                "data": table_dict
                }
            db.tables.insert_one(document)

        print("Tables successfully stored in MongoDB ")
def extract_tables_from_mongodb():
    tables_cursor = db.tables.find({})
    tables = list(tables_cursor)  
    df_list=[]
    if not tables:
        print("No tables found in MongoDB.")
        return
    for table in tables:
        table_name = table.get("table_name", "Unnamed_Table")
        table_data = table.get("data", [])

        if not table_data:
            print(f"Table {table_name} has no data.")
            continue

            
        df = pd.DataFrame(table_data)

            
        if any("_" in col for col in df.columns):
                
            multiindex_columns = [tuple(col.split("_")) for col in df.columns]
            df.columns = pd.MultiIndex.from_tuples(multiindex_columns)
        df_list.append(df)    
        print(table_name)
        print(df)
        if(table_name=="Table_3"):
            df_list[0]=pd.concat([df_list[0],df_list[1]],axis=1)
            df_list[1]=df
        
    return df_list
def write_table_into_excel(df_list):
    file_name = "extracted_tables_multiindex.xlsx"
    try:
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            for i, df in enumerate(df_list):
                
                if(i==0):
                    df.to_excel(writer, sheet_name=f"table {i+1}", index=True,merge_cells=True)
                else:
                    df.to_excel(writer, sheet_name=f"table {i+1}", index=False)
        print(f"Excel file {file_name} written successfully.")
    except Exception as e:
        print(f"Error while writing Excel file: {e}")

def remove_index_from_excel(file_name):
    try:
        wb = load_workbook(file_name)
        
        
        ws = wb[wb.sheetnames[0]]
            
            
        first_row = [cell.value for cell in ws[1]]  
        if first_row[0] in (None, "", "Unnamed: 0", "index"):  
            ws.insert_cols(1)  
            ws.delete_cols(2)  
            
            
        for row in range(ws.max_row, 0, -1):  
            if all(cell.value is None for cell in ws[row]):
                ws.delete_rows(row)
            
            
            
        merge_columns = [1, 2, 3, 4, 5,6]  
        row=1  
        for col in merge_columns:
            current_cell = ws.cell(row=row, column=col)
                    
            if current_cell.value :  
                        
                ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)
                current_cell.value = f"{current_cell.value}"  
                    
        for col in range(1, ws.max_column + 1):  
                max_length = 0
                column = ws.column_dimensions[openpyxl.utils.get_column_letter(col)] # type: ignore
                for row in range(1, ws.max_row + 1):  
                    try:
                        cell_value = str(ws.cell(row=row, column=col).value)
                        max_length = max(max_length, len(cell_value))
                    except:
                        pass
                adjusted_width = (max_length + 2)  
                column.width = adjusted_width  
                
        for row in range(1, ws.max_row + 1):  
                ws.row_dimensions[row].height = 30
        ws=wb[wb.sheetnames[1]] 
        for col in range(1, ws.max_column + 1):  
                max_length = 0
                column = ws.column_dimensions[openpyxl.utils.get_column_letter(col)] # type: ignore
                for row in range(1, ws.max_row + 1):  
                    try:
                        cell_value = str(ws.cell(row=row, column=col).value)
                        max_length = max(max_length, len(cell_value))
                    except:
                        pass
                adjusted_width = (max_length + 2)  
                column.width = adjusted_width  
                
        for row in range(1, ws.max_row + 1):  
                ws.row_dimensions[row].height = 30 
        wb.save(file_name)
        print(f"File processed and saved: {file_name}")
    
    except Exception as e:
        print(f"Error while modifying Excel file: {e}")
def write_table_into_excel_and_remove_index(df_list):
    
    write_table_into_excel(df_list)
    
    
    remove_index_from_excel("extracted_tables_multiindex.xlsx")
def fetch_data_from_mongodb():
    cursor = db.properties.find({}, {  
        "_id": 0,  
        "Style": 1,
        "Style number": 1,
        "Brand": 1,
        "Sizes": 1,
        "Commodity": 1,
        "E-mail": 1,
        "Care Address": 1,
        "Main image":1
    })
    return list(cursor) 

def generate_pdf(properties, output_file, df1):
    custom_page_size = (750, 800)  
    c = canvas.Canvas(output_file, pagesize=custom_page_size)
    width, height = custom_page_size

    
    c.setFont("Helvetica-Bold", 16)
    title = "Costing Sheet"
    title_width = c.stringWidth(title, "Helvetica-Bold", 16)
    c.drawString((width - title_width) / 2, height - 50, title)

    
    c.setFont("Helvetica", 12)
    y_position = height - 100

    def draw_key_value(key, value):
        nonlocal y_position
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y_position, f"{key}:")
        c.setFont("Helvetica", 12)
        c.drawString(150, y_position, value)
        y_position -= 30

    fields = properties[0]
    draw_key_value("Style number", fields["Style number"])
    draw_key_value("Style", fields["Style"])
    draw_key_value("Brand", fields["Brand"])
    draw_key_value("Sizes", fields["Sizes"])
    draw_key_value("Commodity", fields["Commodity"])
    draw_key_value("E-mail", fields["E-mail"])
    draw_key_value("Care Address", fields["Care Address"])

    
    main_image = Image.open(io.BytesIO(properties[0]["Main image"]))
    main_image.save("main_image.jpg")
    image_path = "main_image.jpg"  
    try:
        c.drawImage(image_path, width - 250, height - 250, width=150, height=150)
    except:
        c.setFillColor(colors.red)
        c.drawString(width - 200, height - 150, "Image placeholder")

    y_position -= 50  

    
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y_position, "Spec Sheet:")
    y_position -= 20  

    
    table_data = [list(df1.columns)] + df1.values.tolist()
    for row in table_data[1:]:  
        if isinstance(row[4], (int, float)):  
            row[4] = round(row[4], 2)

    total_value = round(df1['Total'].sum(), 2)
    table_data.append(["Total", "", "", "", str(total_value)])

    
    table = Table(table_data, colWidths=[150, 250, 100, 100, 100])  
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('SPAN', (0, -1), (-2, -1)),  
        ('ALIGN', (0, -1), (-2, -1), 'CENTER'),  
    ]))

    
    table_width, table_height = table.wrap(width, height)
    table_x = (width - table_width) / 2
    table_y = y_position - table_height

    table.wrapOn(c, width, height)
    table.drawOn(c, table_x, table_y)

    c.save()
    print("PDF file is created.")

def download_pdf_attachment(email_user, email_password, folder="inbox"):
    
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(email_user, email_password)
    mail.select(folder)  

    
    status, messages = mail.search(None, 'UNSEEN')
    
    
    email_ids = messages[0].split()

    if not email_ids:
        print("No new emails.")
        return None

    
    latest_email_id = email_ids[-1]
    status, msg_data = mail.fetch(latest_email_id, "(RFC822)")
    
    for response_part in msg_data:
        if isinstance(response_part, tuple):
            msg = email.message_from_bytes(response_part[1])

            
            subject, encoding = decode_header(msg["Subject"])[0]
            if isinstance(subject, bytes):
                subject = subject.decode(encoding if encoding else 'utf-8')

            from_ = msg.get("From")
            print(f"Subject: {subject}")
            print(f"From: {from_}")
            
            
            if msg.is_multipart():
                for part in msg.walk():
                    content_disposition = str(part.get("Content-Disposition"))
                    if "attachment" in content_disposition:
                        
                        filename = part.get_filename()
                        if filename.endswith(".pdf"):  
                            file_path = os.path.join("downloads", filename)
                            if not os.path.isdir("downloads"):
                                os.makedirs("downloads")
                            with open(file_path, "wb") as f:
                                f.write(part.get_payload(decode=True))
                            print(f"Downloaded PDF: {file_path}")
                            return file_path
    return None
def send_email_with_attachments(sender_email, receiver_email, subject, body, smtp_server, smtp_port, smtp_user, smtp_password, pdf_file, excel_file):
    msg = MIMEMultipart()  
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject

    
    msg.attach(MIMEText(body, 'plain'))

    
    with open(pdf_file, 'rb') as pdf:
        pdf_attachment = MIMEApplication(pdf.read(), _subtype='pdf')
        pdf_attachment.add_header('Content-Disposition', 'attachment', filename=os.path.basename(pdf_file))
        msg.attach(pdf_attachment)

    
    with open(excel_file, 'rb') as excel:
        excel_attachment = MIMEApplication(excel.read(), _subtype='octet-stream')
        excel_attachment.add_header('Content-Disposition', 'attachment', filename=os.path.basename(excel_file))
        msg.attach(excel_attachment)

    try:
        
        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
        print("Email sent successfully!")
    except Exception as e:
        print(f"Error sending email: {e}")




#Start execution from this line
email_user = "g92165063@gmail.com"
email_password = "dhkw qyxg ytwm ursb"  

input_pdf_path = download_pdf_attachment(email_user, email_password)

if input_pdf_path:
    
    properties1 = properties(input_pdf_path)
    tables = Tables(input_pdf_path)
    extract_sample_images_and_store_into_mongodb(input_pdf_path)
    storeintomongodb(properties1, tables)
    df_list = extract_tables_from_mongodb()
    df_list = df_list[:-1]

    
    list1 = []
    for i in range(len(df_list[1])):
        list1.append(float(input(f"Enter a per rate value {i+1}: ")))
    df_list[1].insert(2, 'per rate', list1)
    write_table_into_excel_and_remove_index(df_list)

    
    df_list[1]['Qty'] = pd.to_numeric(df_list[1]['Qty'], errors='coerce')
    df_list[1]['per rate'] = pd.to_numeric(df_list[1]['per rate'], errors='coerce')
    df_list[1]['Total'] = df_list[1]['Qty'] * df_list[1]['per rate']
    df_list[1]['Total'] = df_list[1]['Total'].fillna(0).astype(float)

    
    

    
    properties2 = fetch_data_from_mongodb()
    output_pdf_file = "mongodb_reportlab.pdf"
    pos=['Placement','Composition','Qty','per rate','Total']
    df_list[1]=df_list[1][pos]
    generate_pdf(properties2, output_pdf_file, df_list[1])

    
    output_excel_file = "extracted_tables_multiindex.xlsx"
    email_subject = "Processed Costing Sheet"
    email_body = "Please find the processed costing sheet in the attached PDF and Excel files."
    attachments = [output_pdf_file, output_excel_file]

    
    recipient_email = "kumaraguru580@gmail.com"
    send_email_with_attachments(
    sender_email="g92165063@gmail.com",
    receiver_email=recipient_email,
    subject="Test Email with Attachments",
    body="This email contains PDF and Excel attachments.",
    smtp_server="smtp.gmail.com",
    smtp_port=465,
    smtp_user="g92165063@gmail.com",
    smtp_password="dhkw qyxg ytwm ursb",
    pdf_file=r"mongodb_reportlab.pdf",  
    excel_file=r"extracted_tables_multiindex.xlsx"  
)
else:
    print("No input PDF found.")